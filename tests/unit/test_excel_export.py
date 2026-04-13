"""Unit tests for ReportEngine Excel/CSV export (§5A2)."""

from __future__ import annotations

import csv
import os
import tempfile
from pathlib import Path

import pytest

from src.core.report_engine import ReportEngine


_CASES = [
    {"test_item": "VCC nominal", "parameter": "VCC", "condition": "T=25°C",
     "expected_value": "3.3V", "test_method": "DMM", "priority": "high"},
    {"test_item": "Max frequency", "parameter": "Freq", "condition": "VDD=3.3V",
     "expected_value": "168MHz", "test_method": "Oscilloscope", "priority": "high"},
    {"test_item": "Standby current", "parameter": "IDD_SLEEP", "condition": "Sleep mode",
     "expected_value": "<2mA", "test_method": "Ammeter", "priority": "medium"},
]


@pytest.mark.unit
class TestExcelExport:
    @pytest.fixture
    def engine(self, tmp_path: Path) -> ReportEngine:
        return ReportEngine(output_dir=str(tmp_path))

    def test_excel_file_created(self, engine: ReportEngine, tmp_path: Path) -> None:
        path = engine.export_test_cases_excel(_CASES, "STM32F407")
        assert path
        assert Path(path).exists()
        assert path.endswith(".xlsx")

    def test_excel_filename_contains_chip_and_date(self, engine: ReportEngine) -> None:
        path = engine.export_test_cases_excel(_CASES, "STM32F407")
        assert "STM32F407" in Path(path).name
        # Date portion: 8 digits YYYYMMDD
        import re
        assert re.search(r"\d{8}", Path(path).name)

    def test_excel_correct_row_count(self, engine: ReportEngine) -> None:
        import openpyxl
        path = engine.export_test_cases_excel(_CASES, "STM32F407")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Header row + 3 data rows
        assert ws.max_row == 4

    def test_excel_header_row_bold(self, engine: ReportEngine) -> None:
        import openpyxl
        path = engine.export_test_cases_excel(_CASES, "STM32F407")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(row=1, column=1).font.bold

    def test_excel_freeze_panes_set(self, engine: ReportEngine) -> None:
        import openpyxl
        path = engine.export_test_cases_excel(_CASES, "STM32F407")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.freeze_panes is not None

    def test_excel_empty_cases_produces_header_only(self, engine: ReportEngine) -> None:
        import openpyxl
        path = engine.export_test_cases_excel([], "EMPTY_CHIP")
        assert path
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.max_row == 1  # Only header

    def test_excel_columns_match_testcase_fields(self, engine: ReportEngine) -> None:
        import openpyxl
        path = engine.export_test_cases_excel(_CASES, "STM32F407")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        # All 6 columns populated
        assert all(h for h in headers)


@pytest.mark.unit
class TestCsvExport:
    @pytest.fixture
    def engine(self, tmp_path: Path) -> ReportEngine:
        return ReportEngine(output_dir=str(tmp_path))

    def test_csv_file_created(self, engine: ReportEngine) -> None:
        path = engine.export_test_cases_csv(_CASES, "STM32F407")
        assert path
        assert Path(path).exists()
        assert path.endswith(".csv")

    def test_csv_has_correct_rows(self, engine: ReportEngine) -> None:
        path = engine.export_test_cases_csv(_CASES, "STM32F407")
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 3

    def test_csv_has_all_required_columns(self, engine: ReportEngine) -> None:
        path = engine.export_test_cases_csv(_CASES, "STM32F407")
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
        for field in ["test_item", "parameter", "condition", "expected_value", "test_method", "priority"]:
            assert field in headers

    def test_csv_empty_cases_produces_header_only(self, engine: ReportEngine) -> None:
        path = engine.export_test_cases_csv([], "EMPTY_CHIP")
        with open(path, newline="") as f:
            content = f.read()
        lines = [l for l in content.splitlines() if l]
        assert len(lines) == 1  # Only header

    def test_output_dir_created_if_missing(self, tmp_path: Path) -> None:
        subdir = str(tmp_path / "nested" / "exports")
        engine = ReportEngine(output_dir=subdir)
        path = engine.export_test_cases_csv(_CASES, "TEST")
        assert Path(path).exists()
